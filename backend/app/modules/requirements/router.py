import io
import uuid
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy import select, func
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.modules.audit.service import audit
from app.modules.audit.model import AuditAction
from app.modules.auth.deps import get_current_user
from app.modules.auth.schema import TokenData

from .lock import assert_category_unlocked, assert_unlocked_for_requirement
from .model import Requirement, RequirementCategory
from .schema import (
    RequirementCategoryCreate, RequirementCategoryRead, RequirementCategoryUpdate,
    RequirementCreate, RequirementRead, RequirementUpdate, UploadSummary,
)

router = APIRouter(prefix="/requirements", tags=["requirements"])

# ── Initial category seed for new projects ──────────────────────────────────
#
# These are *data*, not behaviour: every project starts with the three
# IEC 62304 regulatory templates wired into a parent chain (USER ← SYSTEM ←
# SOFTWARE). Once the rows exist, runtime behaviour reads from the DB —
# category name, label, color, sort_order, readable_id_prefix, parent_id are
# all fully user-customisable per project.
_INITIAL_CATEGORIES = [
    {"name": "USER",     "label": "User Requirements",     "color": "#1565c0", "sort_order": 0, "readable_id_prefix": "URQ", "parent_name": None},
    {"name": "SYSTEM",   "label": "System Requirements",   "color": "#6a1b9a", "sort_order": 1, "readable_id_prefix": "SYS", "parent_name": "USER"},
    {"name": "SOFTWARE", "label": "Software Requirements", "color": "#1b5e20", "sort_order": 2, "readable_id_prefix": "SWR", "parent_name": "SYSTEM"},
]


async def _ensure_builtins(db: AsyncSession, project_id: uuid.UUID) -> None:
    """Seed initial categories on first use. Idempotent — pre-existing rows
    are kept. Two passes: insert the rows first, then wire parent_id by name."""
    for bc in _INITIAL_CATEGORIES:
        stmt = pg_insert(RequirementCategory).values(
            id=uuid.uuid4(),
            project_id=project_id,
            name=bc["name"],
            label=bc["label"],
            color=bc["color"],
            is_builtin=True,
            sort_order=bc["sort_order"],
            readable_id_prefix=bc["readable_id_prefix"],
            parent_id=None,
        ).on_conflict_do_nothing(constraint="uq_req_category_project_name")
        await db.execute(stmt)
    await db.flush()
    # Wire parents by name lookup, only if currently null (don't clobber user edits).
    cats = (await db.execute(
        select(RequirementCategory).where(RequirementCategory.project_id == project_id)
    )).scalars().all()
    by_name = {c.name: c for c in cats}
    for bc in _INITIAL_CATEGORIES:
        if not bc["parent_name"]:
            continue
        child = by_name.get(bc["name"])
        parent = by_name.get(bc["parent_name"])
        if child and parent and child.parent_id is None:
            child.parent_id = parent.id


# ── Readable ID generation ────────────────────────────────────────────────────

def _prefix_fallback(category_name: str) -> str:
    """Used when a category has no readable_id_prefix set yet. Returns the
    uppercased first 3 letters of the name — keeps IDs deterministic."""
    return category_name.strip().upper()[:3] or "REQ"


async def _next_readable_id(db: AsyncSession, project_id: uuid.UUID, category: RequirementCategory) -> str:
    prefix = (category.readable_id_prefix or _prefix_fallback(category.name)).upper()
    rows = (await db.execute(
        select(Requirement.readable_id).where(
            Requirement.project_id == project_id,
            Requirement.readable_id.like(f"{prefix}-%"),
        )
    )).scalars().all()
    max_n = 0
    for rid in rows:
        try:
            n = int(rid.split("-", 1)[1])
            if n > max_n:
                max_n = n
        except (ValueError, IndexError):
            pass
    return f"{prefix}-{max_n + 1:03d}"


# ── Hierarchy enforcement (driven by RequirementCategory.parent_id) ───────────

async def _validate_hierarchy(
    db: AsyncSession,
    category: RequirementCategory,
    parent_id: uuid.UUID | None,
) -> None:
    """Enforce the parent-of-requirement rule using *category metadata*, not
    hardcoded type names. Rules derived from the category row:

    - Category has no parent (top level, e.g. USER) → requirement must have no parent.
    - Category has parent X → requirement's parent must exist AND be of type X.
    """
    if category.parent_id is None:
        # Top-level category: requirement cannot have a parent.
        if parent_id is not None:
            raise HTTPException(
                400,
                f"'{category.name}' is a top-level category — its requirements cannot have a parent",
            )
        return

    if parent_id is None:
        # Category has a parent declared but the requirement doesn't.
        parent_cat = await db.get(RequirementCategory, category.parent_id)
        parent_name = parent_cat.name if parent_cat else "?"
        raise HTTPException(
            400,
            f"'{category.name}' requirements must have a parent of type '{parent_name}'",
        )

    parent_req = await db.get(Requirement, parent_id)
    if not parent_req:
        raise HTTPException(400, f"Parent requirement {parent_id} not found")
    parent_cat = await db.get(RequirementCategory, category.parent_id)
    if parent_cat and parent_req.type != parent_cat.name:
        raise HTTPException(
            400,
            f"Parent must be a '{parent_cat.name}' requirement, but {parent_req.readable_id} is '{parent_req.type}'",
        )


# ── Category endpoints ────────────────────────────────────────────────────────

@router.get("/categories", response_model=list[RequirementCategoryRead])
async def list_categories(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    await _ensure_builtins(db, project_id)
    await db.commit()
    cats = (
        await db.execute(
            select(RequirementCategory)
            .where(RequirementCategory.project_id == project_id)
            .order_by(RequirementCategory.sort_order, RequirementCategory.name)
        )
    ).scalars().all()
    return cats


@router.post("/categories", response_model=RequirementCategoryRead, status_code=201)
async def create_category(
    body: RequirementCategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    # Note: categories are project-level schema (type taxonomy), not per-category
    # data — no lock check. Mutating category metadata while a baseline is
    # approved is allowed; locking only the data preserves the snapshot's
    # semantic safety.
    existing = (
        await db.execute(
            select(RequirementCategory).where(
                RequirementCategory.project_id == body.project_id,
                RequirementCategory.name == body.name,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(409, f"A category named '{body.name}' already exists for this project")

    # Validate parent belongs to same project
    if body.parent_id:
        parent_cat = (
            await db.execute(
                select(RequirementCategory).where(
                    RequirementCategory.id == body.parent_id,
                    RequirementCategory.project_id == body.project_id,
                )
            )
        ).scalar_one_or_none()
        if not parent_cat:
            raise HTTPException(400, "Parent category not found in this project")

    max_order = (
        await db.execute(
            select(RequirementCategory.sort_order)
            .where(RequirementCategory.project_id == body.project_id)
            .order_by(RequirementCategory.sort_order.desc())
            .limit(1)
        )
    ).scalar_one_or_none() or 2

    cat = RequirementCategory(
        project_id=body.project_id,
        name=body.name,
        label=body.label,
        color=body.color,
        is_builtin=False,
        sort_order=max_order + 1,
        # Pick the user-supplied prefix, falling back to the first 3 letters
        # of the (normalised) name so every category has a usable prefix.
        readable_id_prefix=body.readable_id_prefix or _prefix_fallback(body.name),
        parent_id=body.parent_id,
    )
    db.add(cat)
    await db.flush()
    await audit(db, "requirement_category", cat.id, AuditAction.CREATE, current_user.user_id, f"name={cat.name}")
    await db.commit()
    await db.refresh(cat)
    return cat


@router.put("/categories/{category_id}", response_model=RequirementCategoryRead)
async def update_category(
    category_id: uuid.UUID,
    body: RequirementCategoryUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    cat = (
        await db.execute(select(RequirementCategory).where(RequirementCategory.id == category_id))
    ).scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "Category not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(cat, k, v)
    await audit(db, "requirement_category", cat.id, AuditAction.UPDATE, current_user.user_id)
    await db.commit()
    await db.refresh(cat)
    return cat


@router.delete("/categories/{category_id}", status_code=204)
async def delete_category(
    category_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    cat = (
        await db.execute(select(RequirementCategory).where(RequirementCategory.id == category_id))
    ).scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "Category not found")

    # Check no requirements use this type
    in_use = (
        await db.execute(
            select(Requirement).where(
                Requirement.project_id == cat.project_id,
                Requirement.type == cat.name,
            ).limit(1)
        )
    ).scalar_one_or_none()
    if in_use:
        raise HTTPException(
            400,
            f"Cannot delete '{cat.label}': {cat.name} requirements still exist for this project",
        )
    await audit(db, "requirement_category", cat.id, AuditAction.DELETE, current_user.user_id, f"name={cat.name}")
    await db.delete(cat)
    await db.commit()


# ── Requirement endpoints ─────────────────────────────────────────────────────

@router.get("/", response_model=list[RequirementRead])
async def list_requirements(
    project_id: uuid.UUID | None = None,
    type: str | None = None,
    needs_review: bool | None = None,
    db: AsyncSession = Depends(get_db),
):
    q = select(Requirement)
    if project_id:
        q = q.where(Requirement.project_id == project_id)
    if type:
        q = q.where(Requirement.type == type.upper())
    if needs_review is not None:
        q = q.where(Requirement.needs_review == needs_review)
    result = await db.execute(q)
    return result.scalars().all()


@router.post("/", response_model=RequirementRead, status_code=201)
async def create_requirement(
    payload: RequirementCreate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    await assert_unlocked_for_requirement(db, payload.project_id, payload.type)

    # Validate type exists for project
    cat = (
        await db.execute(
            select(RequirementCategory).where(
                RequirementCategory.project_id == payload.project_id,
                RequirementCategory.name == payload.type.upper(),
            )
        )
    ).scalar_one_or_none()
    if not cat:
        raise HTTPException(400, f"Requirement type '{payload.type}' is not defined for this project")

    await _validate_hierarchy(db, cat, payload.parent_id)
    rid = await _next_readable_id(db, payload.project_id, cat)
    req = Requirement(**{**payload.model_dump(), "type": payload.type.upper(), "readable_id": rid})
    db.add(req)
    await db.flush()
    await audit(db, "requirement", req.id, AuditAction.CREATE, current_user.user_id, f"{rid} {req.type}")
    await db.commit()
    await db.refresh(req)
    return req


@router.get("/{req_id}", response_model=RequirementRead)
async def get_requirement(req_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    req = await db.get(Requirement, req_id)
    if not req:
        raise HTTPException(404, detail="Requirement not found")
    return req


@router.put("/{req_id}", response_model=RequirementRead)
async def update_requirement(
    req_id: uuid.UUID,
    payload: RequirementUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    req = await db.get(Requirement, req_id)
    if not req:
        raise HTTPException(404, detail="Requirement not found")
    await assert_unlocked_for_requirement(db, req.project_id, req.type)

    changed_fields = set(payload.model_dump(exclude_unset=True).keys())
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(req, k, v)
    await audit(
        db, "requirement", req.id, AuditAction.UPDATE, current_user.user_id,
        f"fields={sorted(changed_fields)}",
    )
    await db.commit()
    await db.refresh(req)

    # Flag linked risks for re-evaluation when title/description change (ISO 14971 §10)
    if changed_fields & {"title", "description"}:
        from app.modules.risks.model import Risk, RiskControl
        from sqlalchemy import update as sql_update

        # ── Risks: flag for re-evaluation (ISO 14971 §10) ──────────────────
        direct_ids = (await db.execute(
            select(Risk.id).where(
                Risk.requirement_id == req_id,
                Risk.status != "CLOSED",
            )
        )).scalars().all()
        via_control_ids = (await db.execute(
            select(RiskControl.risk_id).where(RiskControl.requirement_id == req_id)
        )).scalars().all()
        all_risk_ids = list(set(list(direct_ids) + list(via_control_ids)))
        if all_risk_ids:
            await db.execute(
                sql_update(Risk)
                .where(Risk.id.in_(all_risk_ids), Risk.status != "CLOSED")
                .values(re_evaluation_required=True, status="RE_EVALUATION_REQUIRED")
            )
            await audit(
                db, "requirement", req.id, AuditAction.UPDATE, current_user.user_id,
                f"cascade: flagged {len(all_risk_ids)} risk(s) for re-evaluation",
            )

        # ── Cross-category descendants: flag for review (IEC 62304 §6.2) ───
        # Walk parent_id chain downward and mark every descendant needs_review.
        descendants = await _collect_descendants(db, req.id)
        if descendants:
            descendant_ids = [d.id for d in descendants]
            await db.execute(
                sql_update(Requirement)
                .where(Requirement.id.in_(descendant_ids))
                .values(
                    needs_review=True,
                    needs_review_reason=f"Ancestor {req.readable_id} ({req.type}) changed — confirm impact",
                )
            )
            await audit(
                db, "requirement", req.id, AuditAction.UPDATE, current_user.user_id,
                f"cascade: flagged {len(descendant_ids)} downstream requirement(s) for review",
            )

        await db.commit()

    return req


async def _collect_descendants(db: AsyncSession, parent_id: uuid.UUID) -> list[Requirement]:
    """Walk the parent_id chain downward from the given requirement, returning
    all transitively-linked descendants (could span categories USER→SYSTEM→SOFTWARE)."""
    out: list[Requirement] = []
    frontier = [parent_id]
    seen: set[uuid.UUID] = set()
    while frontier:
        children = (await db.execute(
            select(Requirement).where(Requirement.parent_id.in_(frontier))
        )).scalars().all()
        new_ids = []
        for c in children:
            if c.id in seen:
                continue
            seen.add(c.id)
            out.append(c)
            new_ids.append(c.id)
        frontier = new_ids
    return out


@router.delete("/{req_id}", status_code=204)
async def delete_requirement(
    req_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    req = await db.get(Requirement, req_id)
    if not req:
        raise HTTPException(404, detail="Requirement not found")
    await assert_unlocked_for_requirement(db, req.project_id, req.type)
    await audit(
        db, "requirement", req.id, AuditAction.DELETE, current_user.user_id,
        f"{req.readable_id} {req.type}",
    )
    await db.delete(req)
    await db.commit()


# ── Cross-category change-impact (B-bonus2) ───────────────────────────────────

@router.get("/{req_id}/impact-preview")
async def preview_change_impact(req_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    """Return the list of downstream requirements (across all categories) that
    would be flagged needs_review if this requirement's title/description
    changes. Frontend uses this to power the pre-edit confirmation modal."""
    req = await db.get(Requirement, req_id)
    if not req:
        raise HTTPException(404, "Requirement not found")
    descendants = await _collect_descendants(db, req.id)
    return {
        "requirement_id": str(req.id),
        "readable_id": req.readable_id,
        "type": req.type,
        "descendants": [
            {
                "id": str(d.id),
                "readable_id": d.readable_id,
                "type": d.type,
                "title": d.title,
                "needs_review": d.needs_review,
            }
            for d in descendants
        ],
        "total": len(descendants),
        "by_type": {
            t: sum(1 for d in descendants if d.type == t)
            for t in sorted({d.type for d in descendants})
        },
    }


@router.post("/{req_id}/acknowledge-review", response_model=RequirementRead)
async def acknowledge_review(
    req_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    """Clear the needs_review flag on a requirement. Records an audit entry
    so QA can see who acknowledged each impacted item."""
    req = await db.get(Requirement, req_id)
    if not req:
        raise HTTPException(404, "Requirement not found")
    if not req.needs_review:
        return req
    prev_reason = req.needs_review_reason
    req.needs_review = False
    req.needs_review_reason = None
    await audit(
        db, "requirement", req.id, AuditAction.UPDATE, current_user.user_id,
        f"acknowledged review impact ({prev_reason})",
    )
    await db.commit()
    await db.refresh(req)
    return req


# Note: a dedicated "list needs-review" endpoint would collide with the
# UUID-typed /{req_id} path. Use `GET /requirements/?needs_review=true` instead.


@router.post("/upload", response_model=UploadSummary)
async def upload_requirements(
    project_id: uuid.UUID = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: TokenData = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, detail="Only .xlsx files are accepted")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(400, detail="Could not parse Excel file")

    ws = wb.active
    raw_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    if not {"type", "title"}.issubset(set(raw_headers)):
        raise HTTPException(400, detail=f"Excel must have columns: type, title. Found: {raw_headers}")

    # Load allowed categories for this project (the upload's "what types are
    # valid" check is just membership in this list).
    await _ensure_builtins(db, project_id)
    cats = (
        await db.execute(
            select(RequirementCategory).where(RequirementCategory.project_id == project_id)
        )
    ).scalars().all()

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {raw_headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        if not data.get("title"):
            continue
        rows.append(data)

    # Per-category lock check: every type touched by the upload must have a
    # DRAFT/IN_REVIEW baseline (or no baseline at all).
    distinct_types = {r.get("type", "").upper() for r in rows if r.get("type")}
    for t in distinct_types:
        await assert_category_unlocked(db, project_id, t)

    # Process rows in the project's configured category sort_order — so a
    # parent-typed row always lands before its child-typed rows, no matter
    # what custom hierarchy the project defined.
    cat_order = {c.name: c.sort_order for c in cats}
    cat_by_name = {c.name: c for c in cats}
    rows.sort(key=lambda r: cat_order.get(r.get("type", "").upper(), 99999))

    existing_result = await db.execute(
        select(Requirement).where(Requirement.project_id == project_id)
    )
    title_to_req: dict[str, Requirement] = {r.title: r for r in existing_result.scalars()}

    added, skipped = [], []

    for row in rows:
        type_str = row.get("type", "").upper()
        title = row.get("title", "")
        description = row.get("description", "") or None
        parent_title = row.get("parent_title", "") or None

        cat = cat_by_name.get(type_str)
        if cat is None:
            skipped.append({"title": title, "reason": f"Unknown type '{type_str}' — define it in project categories first"})
            continue

        if title in title_to_req:
            skipped.append({"title": title, "reason": "Duplicate title in project"})
            continue

        parent_id = None
        if parent_title:
            if parent_title not in title_to_req:
                skipped.append({"title": title, "reason": f"Parent requirement '{parent_title}' not found"})
                continue
            parent_id = title_to_req[parent_title].id

        rid = await _next_readable_id(db, project_id, cat)
        new_req = Requirement(
            project_id=project_id, type=type_str, readable_id=rid,
            title=title, description=description, parent_id=parent_id,
        )
        db.add(new_req)
        await db.flush()
        await audit(
            db, "requirement", new_req.id, AuditAction.CREATE, current_user.user_id,
            f"bulk-upload {file.filename}: {rid} {type_str}",
        )
        title_to_req[title] = new_req
        added.append({"title": title, "type": type_str})

    # Bulk-event entry against the project so audit log shows a single upload event
    await audit(
        db, "requirement_upload", project_id, AuditAction.CREATE, current_user.user_id,
        f"file={file.filename} added={len(added)} skipped={len(skipped)}",
    )
    await db.commit()
    return UploadSummary(total_added=len(added), total_skipped=len(skipped), added=added, skipped=skipped)
